#!/usr/bin/env python3
"""
Export AI Foundry Model Catalog and Azure ML Registry Models to Excel

This script fetches all available models from Azure AI Foundry (Model Catalog)
and Azure ML Registries, then exports them to an Excel file with details about each model.
"""

import os
import sys
from datetime import datetime
from typing import List, Dict, Any

from azure.mgmt.cognitiveservices import CognitiveServicesManagementClient
from azure.ai.ml import MLClient
from azure.identity import DefaultAzureCredential
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv


def get_management_client() -> tuple[CognitiveServicesManagementClient, str]:
    """
    Create and return a CognitiveServicesManagementClient for accessing Azure AI Foundry.
    
    Returns:
        tuple: (CognitiveServicesManagementClient, location)
    """
    load_dotenv()
    
    subscription_id = os.getenv("AZURE_SUBSCRIPTION_ID")
    location = os.getenv("AZURE_LOCATION")
    
    if not all([subscription_id, location]):
        print("Error: Missing required environment variables.")
        print("Please set AZURE_SUBSCRIPTION_ID and AZURE_LOCATION")
        print("You can copy .env.example to .env and fill in your values.")
        sys.exit(1)
    
    try:
        credential = DefaultAzureCredential()
        client = CognitiveServicesManagementClient(
            credential=credential,
            subscription_id=subscription_id
        )
        return client, location
    except Exception as e:
        print(f"Error creating management client: {e}")
        sys.exit(1)


def fetch_models(client: CognitiveServicesManagementClient, location: str) -> List[Dict[str, Any]]:
    """
    Fetch all models from the AI Foundry catalog using the Account Management API.
    
    Args:
        client: The CognitiveServicesManagementClient instance
        location: Azure region location
        
    Returns:
        List of model dictionaries with their details
    """
    print(f"Fetching models from AI Foundry catalog in location '{location}'...")
    models_data = []
    
    try:
        models = client.models.list(location=location)
        
        for model in models:
            # Extract model details from the Account Management API response
            model_details = model.model if hasattr(model, 'model') else None
            
            if model_details:
                model_info = {
                    "Source": "AI Foundry Catalog",
                    "Name": model_details.name if hasattr(model_details, 'name') else "N/A",
                    "Version": model_details.version if hasattr(model_details, 'version') else "N/A",
                    "Description": model.description if hasattr(model, 'description') and model.description else "N/A",
                    "Format": model_details.format if hasattr(model_details, 'format') else "N/A",
                    "Kind": model.kind if hasattr(model, 'kind') else "N/A",
                    "SKU": model.sku_name if hasattr(model, 'sku_name') else "N/A",
                    "Lifecycle Status": model_details.lifecycle_status if hasattr(model_details, 'lifecycle_status') else "N/A",
                    "Max Capacity": model_details.max_capacity if hasattr(model_details, 'max_capacity') else "N/A",
                }
                
                # Add system data if available
                if hasattr(model_details, 'system_data') and model_details.system_data:
                    system_data = model_details.system_data
                    model_info["Created Date"] = system_data.created_at.strftime("%Y-%m-%d %H:%M:%S") if hasattr(system_data, 'created_at') and system_data.created_at else "N/A"
                    model_info["Created By"] = system_data.created_by if hasattr(system_data, 'created_by') else "N/A"
                    model_info["Last Modified Date"] = system_data.last_modified_at.strftime("%Y-%m-%d %H:%M:%S") if hasattr(system_data, 'last_modified_at') and system_data.last_modified_at else "N/A"
                    model_info["Last Modified By"] = system_data.last_modified_by if hasattr(system_data, 'last_modified_by') else "N/A"
                else:
                    model_info["Created Date"] = "N/A"
                    model_info["Created By"] = "N/A"
                    model_info["Last Modified Date"] = "N/A"
                    model_info["Last Modified By"] = "N/A"
                
                models_data.append(model_info)
            
        print(f"Found {len(models_data)} models")
        return models_data
        
    except Exception as e:
        print(f"Error fetching models: {e}")
        import traceback
        traceback.print_exc()
        return []


def fetch_registry_models(credential: DefaultAzureCredential, registry_names: List[str]) -> List[Dict[str, Any]]:
    """
    Fetch all models from Azure ML Registries.
    
    Args:
        credential: Azure credential object
        registry_names: List of registry names to fetch models from (e.g., ['azureml', 'azureml-meta'])
        
    Returns:
        List of model dictionaries with their details
    """
    all_models_data = []
    
    for registry_name in registry_names:
        print(f"Fetching models from Azure ML Registry '{registry_name}'...")
        try:
            # Create MLClient for the registry
            ml_client = MLClient(credential=credential, registry_name=registry_name)
            
            # List all models in the registry
            models = ml_client.models.list()
            
            count = 0
            for model in models:
                model_info = {
                    "Source": f"Azure ML Registry ({registry_name})",
                    "Name": model.name if hasattr(model, 'name') else "N/A",
                    "Version": str(model.version) if hasattr(model, 'version') else "N/A",
                    "Description": model.description if hasattr(model, 'description') and model.description else "N/A",
                    "Format": model.type if hasattr(model, 'type') else "N/A",
                    "Kind": "N/A",
                    "SKU": "N/A",
                    "Lifecycle Status": model.stage if hasattr(model, 'stage') and model.stage else "N/A",
                    "Max Capacity": "N/A",
                }
                
                # Add tags as additional info if available
                if hasattr(model, 'tags') and model.tags:
                    tags_str = ", ".join([f"{k}={v}" for k, v in model.tags.items()]) if isinstance(model.tags, dict) else str(model.tags)
                    if model_info["Description"] == "N/A":
                        model_info["Description"] = f"Tags: {tags_str}"
                    else:
                        model_info["Description"] += f" | Tags: {tags_str}"
                
                # Add creation metadata if available
                if hasattr(model, 'creation_context') and model.creation_context:
                    creation_context = model.creation_context
                    model_info["Created Date"] = creation_context.created_at.strftime("%Y-%m-%d %H:%M:%S") if hasattr(creation_context, 'created_at') and creation_context.created_at else "N/A"
                    model_info["Created By"] = creation_context.created_by if hasattr(creation_context, 'created_by') else "N/A"
                    model_info["Last Modified Date"] = creation_context.last_modified_at.strftime("%Y-%m-%d %H:%M:%S") if hasattr(creation_context, 'last_modified_at') and creation_context.last_modified_at else "N/A"
                    model_info["Last Modified By"] = creation_context.last_modified_by if hasattr(creation_context, 'last_modified_by') else "N/A"
                else:
                    model_info["Created Date"] = "N/A"
                    model_info["Created By"] = "N/A"
                    model_info["Last Modified Date"] = "N/A"
                    model_info["Last Modified By"] = "N/A"
                
                all_models_data.append(model_info)
                count += 1
            
            print(f"Found {count} models in registry '{registry_name}'")
            
        except Exception as e:
            print(f"Error fetching models from registry '{registry_name}': {e}")
            import traceback
            traceback.print_exc()
    
    return all_models_data


def export_to_excel(models_data: List[Dict[str, Any]], output_file: str):
    """
    Export models data to an Excel file with formatting.
    
    Args:
        models_data: List of model dictionaries
        output_file: Path to output Excel file
    """
    print(f"Exporting {len(models_data)} models to Excel...")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "AI Foundry Models"
    
    # Define headers
    headers = ["Source", "Name", "Version", "Description", "Format", "Kind", "SKU", "Lifecycle Status", "Max Capacity", "Created Date", "Created By", "Last Modified Date", "Last Modified By"]
    
    # Style for headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Write data
    for row_num, model in enumerate(models_data, 2):
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = model.get(header, "N/A")
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    
    # Auto-adjust column widths
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        for cell in ws[column_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except (TypeError, AttributeError):
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze the header row
    ws.freeze_panes = "A2"
    
    # Save the workbook
    wb.save(output_file)
    print(f"Excel file saved to: {output_file}")


def main():
    """Main execution function."""
    print("=" * 60)
    print("AI Foundry Models to Excel Exporter")
    print("=" * 60)
    print()
    
    # Get management client and location
    client, location = get_management_client()
    
    # Fetch models from AI Foundry catalog
    models_data = fetch_models(client, location)
    
    # Fetch models from Azure ML Registries
    load_dotenv()
    registry_names_str = os.getenv("AZURE_ML_REGISTRY_NAMES", "azureml,azureml-meta,azureml-cohere,azureml-mistral,azureml-xai,HuggingFace,azureml-nvidia")
    registry_names = [name.strip() for name in registry_names_str.split(",") if name.strip()]
    
    if registry_names:
        print()
        print(f"Configured Azure ML Registries: {', '.join(registry_names)}")
        credential = DefaultAzureCredential()
        registry_models = fetch_registry_models(credential, registry_names)
        models_data.extend(registry_models)
    
    if not models_data:
        print("No models found or error occurred.")
        sys.exit(1)
    
    # Generate output filename with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f"ai_foundry_models_{timestamp}.xlsx"
    
    # Export to Excel
    export_to_excel(models_data, output_file)
    
    print()
    print("=" * 60)
    print(f"Total models exported: {len(models_data)}")
    print("Export completed successfully!")
    print("=" * 60)


if __name__ == "__main__":
    main()
